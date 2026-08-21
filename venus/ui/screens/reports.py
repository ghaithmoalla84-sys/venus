# -*- coding: utf-8 -*-
"""
شاشة التقارير - Venus Coffee
تقارير متعددة التبويبات: المبيعات، الأرباح التقديرية، المخزون، الديون، حركة النقدية
"""

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QLabel, QLineEdit,
     QPushButton, QTableWidget, QTableWidgetItem, QTabWidget, QMessageBox,
     QDateEdit, QHeaderView, QGroupBox, QStyledItemDelegate, QStyle, QApplication,
     QComboBox
)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QDoubleValidator, QColor
from datetime import datetime

from venus.core.database import get_conn, today_str, yesterday_str, patch_db_path
from venus.core.events import app_events
from venus.ui.widgets.loading_overlay import LoadingOverlay
from venus.ui.widgets.searchable_table import SearchableTable
from venus.ui.widgets.sales_chart_widget import SalesChartWidget
from venus.ui.styles import (
    Colors, FontSizes, Spacing, BorderRadius,
    title_label_style, tab_style, group_box_style, table_style,
    primary_button_style, summary_label_style,
    date_edit_style
)
from venus.utils.currency import fmt, fmt_syp, fmt_usd
from venus.utils.logger import setup_logger
from venus.utils.overdue import get_overdue_debts
logger = setup_logger()

class NumericDelegate(QStyledItemDelegate):
    """ديلجيت لتقييد الإدخال بالأرقام فقط في الجداول"""
    def createEditor(self, parent, option, index):
        editor = QLineEdit(parent)
        editor.setValidator(QDoubleValidator())
        return editor

class ReportsScreen(QWidget):
    """شاشة التقارير المتعددة"""

    def __init__(self):
        super().__init__()
        self.setLayoutDirection(Qt.RightToLeft)
        self.exchange_rate = None
        self.init_ui()
        self._load_suppliers_combo()
        self.load_exchange_rate()

        app_events.data_changed.connect(self._on_app_data_changed)

    def _run_with_loading(self, button, func):
        button.setEnabled(False)
        self.loading_overlay.start()
        QApplication.processEvents()
        try:
            func()
        finally:
            self.loading_overlay.stop()
            button.setEnabled(True)

    # ─────────────────────────────────────────────
    # بناء الواجهة الرئيسية
    # ─────────────────────────────────────────────
    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(30, 30, 30, 30)

        title = QLabel("📊 التقارير - Venus Coffee")
        title.setStyleSheet(title_label_style(font_size=FontSizes.XL6, color=Colors.DARK))
        title.setAlignment(Qt.AlignRight)
        main_layout.addWidget(title)

        tabs = QTabWidget()
        tabs.setLayoutDirection(Qt.RightToLeft)
        tabs.setStyleSheet(tab_style())
        self.tabs = tabs
        main_layout.addWidget(tabs)

        self.sales_tab = QWidget()
        self.profit_tab = QWidget()
        self.inventory_tab = QWidget()
        self.buy_list_tab = QWidget()
        self.debts_tab = QWidget()
        self.cash_tab = QWidget()
        self.suppliers_tab = QWidget()
        self.overdue_tab = QWidget()
        self.best_suppliers_tab = QWidget()

        tabs.addTab(self.sales_tab, "🛒 المبيعات")
        tabs.addTab(self.profit_tab, "📈 الأرباح التقديرية")
        tabs.addTab(self.inventory_tab, "📦 المخزون")
        tabs.addTab(self.buy_list_tab, "🛒 ما يجب شراؤه الآن")
        tabs.addTab(self.debts_tab, "💳 الديون")
        tabs.addTab(self.cash_tab, "💰 حركة النقدية")
        tabs.addTab(self.suppliers_tab, "🏭 الموردون")
        tabs.addTab(self.overdue_tab, "⏰ الديون المتأخرة")
        tabs.addTab(self.best_suppliers_tab, "🏆 أفضل الموردين")

        self.comparison_tab = QWidget()
        tabs.addTab(self.comparison_tab, "📊 مقارنة الفترات")

        self.tax_report_tab = QWidget()
        tabs.addTab(self.tax_report_tab, "📑 التقرير الضريبي")

        self.build_sales_tab()
        self.build_profit_tab()
        self.build_inventory_tab()
        self.build_buy_list_tab()
        self.build_debts_tab()
        self.build_cash_tab()
        self.build_suppliers_tab()
        self.build_overdue_tab()
        self.build_best_suppliers_tab()
        self.build_comparison_tab()
        self.build_tax_report_tab()

        self.loading_overlay = LoadingOverlay(self)
        main_layout.addWidget(self.loading_overlay)

    # ─────────────────────────────────────────────
    # تبويب المبيعات
    # ─────────────────────────────────────────────
    def build_sales_tab(self):
        layout = QVBoxLayout(self.sales_tab)
        layout.setSpacing(20)
        layout.setContentsMargins(20, 20, 20, 20)

        filters = QWidget()
        f_layout = QHBoxLayout(filters)
        f_layout.setContentsMargins(0, 0, 0, 0)

        f_layout.addWidget(QLabel("📅 من:"))
        self.sales_from = QDateEdit()
        self.sales_from.setDate(QDate.currentDate().addMonths(-1))
        self.sales_from.setCalendarPopup(True)
        self.sales_from.setStyleSheet(self._date_style())
        f_layout.addWidget(self.sales_from)

        f_layout.addWidget(QLabel("📅 إلى:"))
        self.sales_to = QDateEdit()
        self.sales_to.setDate(QDate.currentDate())
        self.sales_to.setCalendarPopup(True)
        self.sales_to.setStyleSheet(self._date_style())
        f_layout.addWidget(self.sales_to)

        show_btn = QPushButton("عرض التقرير")
        show_btn.setIcon(QApplication.style().standardIcon(QStyle.SP_FileDialogDetailedView))
        show_btn.setCursor(Qt.PointingHandCursor)
        show_btn.setStyleSheet(self._btn_style("#3498db", "#2980b9"))
        show_btn.clicked.connect(lambda: self._run_with_loading(show_btn, self.load_sales_report))
        f_layout.addWidget(show_btn)

        btn_excel = QPushButton("📊 تصدير Excel")
        btn_excel.setIcon(QApplication.style().standardIcon(QStyle.SP_ArrowUp))
        btn_excel.setCursor(Qt.PointingHandCursor)
        btn_excel.setStyleSheet(self._btn_style("#3498db", "#2980b9"))
        btn_excel.clicked.connect(lambda: self.export_table_to_excel(
            self.sales_table, "تقرير_المبيعات.xlsx", "المبيعات",
            ["المجموعة", "عدد العمليات", "إجمالي المبلغ"]
        ))
        f_layout.addWidget(btn_excel)

        btn_pdf = QPushButton("📄 تصدير PDF")
        btn_pdf.setIcon(QApplication.style().standardIcon(QStyle.SP_ArrowUp))
        btn_pdf.setCursor(Qt.PointingHandCursor)
        btn_pdf.setStyleSheet(self._btn_style("#3498db", "#2980b9"))
        btn_pdf.clicked.connect(lambda: self.export_table_to_pdf(
            self.sales_table, "تقرير_المبيعات.pdf", "تقرير المبيعات",
            ["المجموعة", "عدد العمليات", "إجمالي المبلغ"]
        ))
        f_layout.addWidget(btn_pdf)

        f_layout.addStretch()
        layout.addWidget(filters)

        self.sales_table = SearchableTable(show_actions=False)
        layout.addWidget(self.sales_table)

        self.sales_total_label = QLabel("الإجمالي العام: 0")
        self.sales_total_label.setStyleSheet(f"""
            font-size: {FontSizes.XL3};
            font-weight: bold;
            color: {Colors.DARK};
            padding: 12px;
            background-color: #eaf2f8;
            border-radius: {BorderRadius.LG};
        """)
        self.sales_total_label.setAlignment(Qt.AlignRight)
        layout.addWidget(self.sales_total_label)

        self.sales_chart_widget = SalesChartWidget()
        layout.addWidget(self.sales_chart_widget)

        self.sales_chart_note = QLabel("تُعرض أعلى 10 مجموعات فقط")
        self.sales_chart_note.setStyleSheet(f"""
            color: {Colors.SECONDARY_TEXT};
            font-size: {FontSizes.SM};
            font-style: italic;
            padding: 4px 8px;
        """)
        self.sales_chart_note.setAlignment(Qt.AlignCenter)
        self.sales_chart_note.hide()
        layout.addWidget(self.sales_chart_note)

    def load_sales_report(self):
        from_date = self.sales_from.date().toString("yyyy-MM-dd")
        to_date = self.sales_to.date().toString("yyyy-MM-dd")
        from_date = from_date.translate(str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789"))
        to_date = to_date.translate(str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789"))
        conn = get_conn()
        try:
            cur = conn.cursor()
            cur.execute("""
                SELECT
                    ج.الاسم AS اسم_المجموعة,
                    COUNT(*) AS عدد_العمليات,
                    SUM(م.المبلغ_الإجمالي) AS إجمالي_المبلغ
                FROM المبيعات_اليومية م
                JOIN المجموعات ج ON م.معرف_المجموعة = ج.معرف
                WHERE date(normalize_date(م.التاريخ)) >= ? AND date(normalize_date(م.التاريخ)) <= ?
                GROUP BY م.معرف_المجموعة
                ORDER BY إجمالي_المبلغ DESC
            """, (from_date, to_date))
            sales_rows = cur.fetchall()

            cur.execute("""
                SELECT DISTINCT العملة FROM المبيعات_اليومية
                WHERE date(normalize_date(التاريخ)) >= ? AND date(normalize_date(التاريخ)) <= ?
            """, (from_date, to_date))
            currencies = [r["العملة"] for r in cur.fetchall()]
            if len(currencies) == 1:
                currency_label = currencies[0] or "ليرة سورية"
            elif len(currencies) > 1:
                currency_label = "عملات متعددة (" + " و".join(currencies) + ")"
            else:
                currency_label = "ليرة سورية"

            rows_data = []
            total = 0.0
            for row in sales_rows:
                amount = row["إجمالي_المبلغ"] or 0
                rows_data.append([
                    row["اسم_المجموعة"] or "",
                    str(row["عدد_العمليات"] or 0),
                    fmt(amount)
                ])
                total += amount

            self.sales_table.set_data(
                ["المجموعة", "عدد العمليات", "إجمالي المبلغ"],
                rows_data
            )

            self.sales_total_label.setText(f"الإجمالي العام: {fmt(total)} {currency_label}")

            chart_data = [
                (row["اسم_المجموعة"] or "", row["إجمالي_المبلغ"] or 0)
                for row in sales_rows
            ]
            if len(chart_data) > 10:
                chart_data = chart_data[:10]
                self.sales_chart_note.show()
            else:
                self.sales_chart_note.hide()
            self.sales_chart_widget.set_data(chart_data)
        except Exception as e:
            logger.error(str(e))
            QMessageBox.critical(self, "خطأ", f"فشل تحميل تقرير المبيعات:\n{str(e)}")
        finally:
            conn.close()

    # ─────────────────────────────────────────────
    # تبويب الأرباح التقديرية
    # ─────────────────────────────────────────────
    def build_profit_tab(self):
        layout = QVBoxLayout(self.profit_tab)
        layout.setSpacing(20)
        layout.setContentsMargins(20, 20, 20, 20)

        filters = QWidget()
        f_layout = QHBoxLayout(filters)
        f_layout.setContentsMargins(0, 0, 0, 0)

        f_layout.addWidget(QLabel("📅 من:"))
        self.profit_from = QDateEdit()
        self.profit_from.setDate(QDate.currentDate().addMonths(-1))
        self.profit_from.setCalendarPopup(True)
        self.profit_from.setStyleSheet(self._date_style())
        f_layout.addWidget(self.profit_from)

        f_layout.addWidget(QLabel("📅 إلى:"))
        self.profit_to = QDateEdit()
        self.profit_to.setDate(QDate.currentDate())
        self.profit_to.setCalendarPopup(True)
        self.profit_to.setStyleSheet(self._date_style())
        f_layout.addWidget(self.profit_to)

        show_btn = QPushButton("حساب الأرباح")
        show_btn.setIcon(QApplication.style().standardIcon(QStyle.SP_FileDialogDetailedView))
        show_btn.setCursor(Qt.PointingHandCursor)
        show_btn.setStyleSheet(self._btn_style("#9b59b6", "#8e44ad"))
        show_btn.clicked.connect(lambda: self._run_with_loading(show_btn, self.load_profit_report))
        f_layout.addWidget(show_btn)

        export_excel_btn = QPushButton("📊 Excel")
        export_excel_btn.setStyleSheet(self._btn_style("#27ae60", "#229954"))
        export_excel_btn.clicked.connect(lambda: self.export_table_to_excel(
            self.profit_table, "تقرير_الأرباح.xlsx", "الأرباح",
            ["المجموعة", "مخزون البداية", "المشتريات", "مخزون النهاية", "التكلفة", "صافي الربح"]
        ))
        f_layout.addWidget(export_excel_btn)

        export_pdf_btn = QPushButton("📄 PDF")
        export_pdf_btn.setStyleSheet(self._btn_style("#e74c3c", "#c0392b"))
        export_pdf_btn.clicked.connect(lambda: self.export_table_to_pdf(
            self.profit_table, "تقرير_الأرباح.pdf", "تقرير الأرباح التقديرية",
            ["المجموعة", "مخزون البداية", "المشتريات", "مخزون النهاية", "التكلفة", "صافي الربح"]
        ))
        f_layout.addWidget(export_pdf_btn)

        f_layout.addStretch()
        layout.addWidget(filters)

        self.profit_table = QTableWidget()
        self.profit_table.setColumnCount(6)
        self.profit_table.setHorizontalHeaderLabels([
            "المجموعة", "مخزون البداية", "المشتريات", "مخزون النهاية", "التكلفة", "صافي الربح"
        ])
        self.profit_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.profit_table.horizontalHeader().setLayoutDirection(Qt.RightToLeft)
        self.profit_table.setStyleSheet(self._table_style("#9b59b6"))
        self.profit_table.setAlternatingRowColors(True)
        layout.addWidget(self.profit_table)

        self.profit_info = QLabel("")
        self.profit_info.setStyleSheet(summary_label_style())
        self.profit_info.setAlignment(Qt.AlignRight)
        layout.addWidget(self.profit_info)

    def _get_audit_inventory_by_group(self, cur, cutoff_date, inclusive=False):
        """جلب قيمة مخزون الجرد لكل مجموعة حسب آخر جرد قبل/عند تاريخ معين."""
        if inclusive:
            date_condition = "date(normalize_date(التاريخ)) <= ?"
        else:
            date_condition = "date(normalize_date(التاريخ)) < ?"

        cur.execute(f"""
            SELECT
                م.معرف_المجموعة,
                SUM(j.الكمية_الفعلي * م.سعر_الشراء_الأخير) AS قيمة_المخزون
            FROM (
                SELECT معرف_المادة_الفرعية, MAX(معرف) AS max_id
                FROM الجرد
                WHERE {date_condition}
                GROUP BY معرف_المادة_الفرعية
            ) آخر_جرد
            JOIN الجرد j ON j.معرف = آخر_جرد.max_id
            JOIN المواد_الفرعية م ON j.معرف_المادة_الفرعية = م.معرف
            GROUP BY م.معرف_المجموعة
        """, (cutoff_date,))
        return {row["معرف_المجموعة"]: (row["قيمة_المخزون"] or 0) for row in cur.fetchall()}

    def load_profit_report(self):
        from_date = self.profit_from.date().toString("yyyy-MM-dd")
        to_date = self.profit_to.date().toString("yyyy-MM-dd")
        from_date = from_date.translate(str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789"))
        to_date = to_date.translate(str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789"))
        conn = get_conn()
        try:
            cur = conn.cursor()

            opening_inventory = self._get_audit_inventory_by_group(cur, from_date, inclusive=False)
            closing_inventory = self._get_audit_inventory_by_group(cur, to_date, inclusive=True)

            cur.execute("""
                SELECT م.معرف_المجموعة, SUM(د.المبلغ_الإجمالي) AS إجمالي_المشتريات
                FROM تفاصيل_الشراء د
                JOIN فواتير_الشراء ف ON د.معرف_الفاتورة = ف.معرف
                JOIN المواد_الفرعية م ON د.معرف_المادة_الفرعية = م.معرف
                WHERE date(normalize_date(ف.التاريخ)) >= ? AND date(normalize_date(ف.التاريخ)) <= ?
                GROUP BY م.معرف_المجموعة
            """, (from_date, to_date))
            purchase_totals = {row["معرف_المجموعة"]: (row["إجمالي_المشتريات"] or 0) for row in cur.fetchall()}

            cur.execute("""
                SELECT معرف_المجموعة, SUM(المبلغ_الإجمالي) AS إجمالي_المبيعات
                FROM المبيعات_اليومية
                WHERE date(normalize_date(التاريخ)) >= ? AND date(normalize_date(التاريخ)) <= ?
                GROUP BY معرف_المجموعة
            """, (from_date, to_date))
            sales_totals = {row["معرف_المجموعة"]: (row["إجمالي_المبيعات"] or 0) for row in cur.fetchall()}

            cur.execute("SELECT معرف, الاسم FROM المجموعات ORDER BY الترتيب, الاسم")
            groups = cur.fetchall()

            report_rows = []
            for g in groups:
                gid = g["معرف"]
                sales = sales_totals.get(gid, 0)
                opening = opening_inventory.get(gid, 0)
                purchases = purchase_totals.get(gid, 0)
                closing = closing_inventory.get(gid, 0)
                cost = opening + purchases - closing
                profit = sales - cost
                report_rows.append({
                    "group_name": g["الاسم"],
                    "opening": opening,
                    "purchases": purchases,
                    "closing": closing,
                    "cost": cost,
                    "profit": profit
                })

            self.profit_table.setRowCount(len(report_rows))
            total_opening = 0.0
            total_purchases = 0.0
            total_closing = 0.0
            total_cost = 0.0
            total_profit = 0.0
            for r, row in enumerate(report_rows):
                self.profit_table.setItem(r, 0, QTableWidgetItem(str(row["group_name"])))
                self.profit_table.setItem(r, 1, QTableWidgetItem(fmt(row['opening'])))
                self.profit_table.setItem(r, 2, QTableWidgetItem(fmt(row['purchases'])))
                self.profit_table.setItem(r, 3, QTableWidgetItem(fmt(row['closing'])))
                self.profit_table.setItem(r, 4, QTableWidgetItem(fmt(row['cost'])))
                profit_text = fmt(row['profit'])
                if row['opening'] == 0 and row['closing'] == 0:
                    profit_text += " (تقديري - بدون جرد)"
                self.profit_table.setItem(r, 5, QTableWidgetItem(profit_text))
                profit_val = row["profit"]
                profit_item = self.profit_table.item(r, 5)
                if profit_val > 0:
                    profit_item.setForeground(QColor("#27ae60"))
                elif profit_val < 0:
                    profit_item.setForeground(QColor("#e74c3c"))
                total_opening += row["opening"]
                total_purchases += row["purchases"]
                total_closing += row["closing"]
                total_cost += row["cost"]
                total_profit += row["profit"]

            self.profit_info.setText(
                f"📅 من: {from_date} | إلى: {to_date} | "
                f"إجمالي المبيعات: {fmt(total_cost + total_profit)} | "
                f"إجمالي التكلفة: {fmt(total_cost)} | "
                f"صافي الربح التقديري: {fmt(total_profit)}"
            )
        except Exception as e:
            logger.error(str(e))
            QMessageBox.critical(self, "خطأ", f"فشل حساب الأرباح:\n{str(e)}")
        finally:
            conn.close()

    # ─────────────────────────────────────────────
    # تبويب المخزون
    # ─────────────────────────────────────────────
    def build_inventory_tab(self):
        layout = QVBoxLayout(self.inventory_tab)
        layout.setSpacing(20)
        layout.setContentsMargins(20, 20, 20, 20)

        header = QWidget()
        h_layout = QHBoxLayout(header)
        h_layout.setContentsMargins(0, 0, 0, 0)

        refresh_btn = QPushButton("تحديث")
        refresh_btn.setIcon(QApplication.style().standardIcon(QStyle.SP_BrowserReload))
        refresh_btn.setCursor(Qt.PointingHandCursor)
        refresh_btn.setStyleSheet(self._btn_style("#27ae60", "#229954"))
        refresh_btn.clicked.connect(lambda: self._run_with_loading(refresh_btn, self.load_inventory))
        h_layout.addWidget(refresh_btn)

        btn_excel = QPushButton("📊 تصدير Excel")
        btn_excel.setIcon(QApplication.style().standardIcon(QStyle.SP_ArrowUp))
        btn_excel.setCursor(Qt.PointingHandCursor)
        btn_excel.setStyleSheet(self._btn_style("#e67e22", "#d35400"))
        btn_excel.clicked.connect(lambda: self.export_table_to_excel(
            self.inventory_table, "المخزون.xlsx", "المخزون",
            ["المادة", "المجموعة", "الوحدة", "الكمية"]
        ))
        h_layout.addWidget(btn_excel)

        btn_pdf = QPushButton("📄 تصدير PDF")
        btn_pdf.setIcon(QApplication.style().standardIcon(QStyle.SP_ArrowUp))
        btn_pdf.setCursor(Qt.PointingHandCursor)
        btn_pdf.setStyleSheet(self._btn_style("#e67e22", "#d35400"))
        btn_pdf.clicked.connect(lambda: self.export_table_to_pdf(
            self.inventory_table, "المخزون.pdf", "تقرير المخزون",
            ["المادة", "المجموعة", "الوحدة", "الكمية"]
        ))
        h_layout.addWidget(btn_pdf)

        h_layout.addStretch()
        layout.addWidget(header)

        self.inventory_table = SearchableTable(show_actions=False)
        layout.addWidget(self.inventory_table)

        self.load_inventory()

    # ─────────────────────────────────────────────
    # تبويب ما يجب شراؤه الآن (مخزون ذكي)
    # ─────────────────────────────────────────────
    def build_buy_list_tab(self):
        layout = QVBoxLayout(self.buy_list_tab)
        layout.setSpacing(20)
        layout.setContentsMargins(20, 20, 20, 20)

        header = QWidget()
        h_layout = QHBoxLayout(header)
        h_layout.setContentsMargins(0, 0, 0, 0)

        refresh_btn = QPushButton("تحديث")
        refresh_btn.setIcon(QApplication.style().standardIcon(QStyle.SP_BrowserReload))
        refresh_btn.setCursor(Qt.PointingHandCursor)
        refresh_btn.setStyleSheet(self._btn_style("#e67e22", "#d35400"))
        refresh_btn.clicked.connect(lambda: self._run_with_loading(refresh_btn, self.load_buy_list))
        h_layout.addWidget(refresh_btn)

        btn_excel = QPushButton("📊 تصدير Excel")
        btn_excel.setIcon(QApplication.style().standardIcon(QStyle.SP_ArrowUp))
        btn_excel.setCursor(Qt.PointingHandCursor)
        btn_excel.setStyleSheet(self._btn_style("#e67e22", "#d35400"))
        btn_excel.clicked.connect(lambda: self.export_table_to_excel(
            self.buy_list_table, "ما_يجب_شراؤه.xlsx", "قائمة الشراء",
            ["المادة", "المجموعة", "الوحدة", "الكمية الحالية", "الحد الأدنى", "الفرق", "مقترح شراء", "معدل استهلاك شهري"]
        ))
        h_layout.addWidget(btn_excel)

        btn_pdf = QPushButton("📄 تصدير PDF")
        btn_pdf.setIcon(QApplication.style().standardIcon(QStyle.SP_ArrowUp))
        btn_pdf.setCursor(Qt.PointingHandCursor)
        btn_pdf.setStyleSheet(self._btn_style("#e67e22", "#d35400"))
        btn_pdf.clicked.connect(lambda: self.export_table_to_pdf(
            self.buy_list_table, "ما_يجب_شراؤه.pdf", "تقرير ما يجب شراؤه الآن",
            ["المادة", "المجموعة", "الوحدة", "الكمية الحالية", "الحد الأدنى", "الفرق", "مقترح شراء", "معدل استهلاك شهري"]
        ))
        h_layout.addWidget(btn_pdf)

        h_layout.addStretch()
        layout.addWidget(header)

        self.buy_list_table = SearchableTable(show_actions=False)
        layout.addWidget(self.buy_list_table)

        self.buy_list_info = QLabel("")
        self.buy_list_info.setStyleSheet(summary_label_style(bg="#fef9e7", border_color="#f9e79f"))
        self.buy_list_info.setAlignment(Qt.AlignRight)
        self.buy_list_info.setWordWrap(True)
        layout.addWidget(self.buy_list_info)

        self.load_buy_list()

    def load_buy_list(self):
        """تحميل تقرير 'ما يجب شراؤه الآن'.

        يُظهر المواد التي الكمية_المتوفرة ≤ الحد_الأدنى و الحد_الأدنى > 0،
        مع معدل الاستهلاك الشهري التقديري القائم على آخر عمليتي جرد متتاليتين.
        """
        from venus.utils.inventory_analytics import get_buy_list

        conn = get_conn()
        try:
            cur = conn.cursor()
            items = get_buy_list(cur)

            headers = [
                "المادة", "المجموعة", "الوحدة", "الكمية الحالية",
                "الحد الأدنى", "الفرق", "مقترح شراء", "معدل استهلاك شهري"
            ]
            rows_data = []
            for item in items:
                rows_data.append([
                    item["name"] or "",
                    item["group_name"] or "",
                    item["unit"] or "",
                    f"{item['current_qty']:,.2f}",
                    f"{item['min_qty']:,.2f}",
                    f"{item['diff']:,.2f}",
                    f"{item['suggested_qty']:,.2f}",
                    item["monthly_consumption"],
                ])

            self.buy_list_table.set_data(headers, rows_data, id_column_index=0)

            if not items:
                self.buy_list_info.setText(
                    "📋 لا توجد مواد تحت الحد الأدنى المحدد حالياً."
                )
            else:
                critical = [i for i in items if i["consumption_reason"] is not None]
                parts = [f"🔍 عدد المواد المطلوبة: {len(items)}"]
                if critical:
                    parts.append(f"⚠️ {len(critical)} مادة لا توفر بيانات جرد كافية لحساب الاستهلاك")
                parts.append(
                    "ملاحظة: معدل الاستهلاك تقديري بناءً على آخر دورة جرد، "
                    "وليس متوسطاً عبر فترة أطول."
                )
                self.buy_list_info.setText(" | ".join(parts))
        except Exception as e:
            logger.error(str(e))
            QMessageBox.critical(self, "خطأ", f"فشل تحميل تقرير ما يجب شراؤه:\n{str(e)}")
        finally:
            conn.close()

    def load_inventory(self):
        conn = get_conn()
        try:
            cur = conn.cursor()
            cur.execute("""
                SELECT
                    م.الاسم AS اسم_المادة,
                    م.الوحدة,
                    ج.الاسم AS اسم_المجموعة,
                    خ.الكمية_المتوفرة,
                    خ.آخر_تحديث
                FROM المخزون خ
                JOIN المواد_الفرعية م ON خ.معرف_المادة_الفرعية = م.معرف
                JOIN المجموعات ج ON م.معرف_المجموعة = ج.معرف
                ORDER BY ج.الاسم, م.الاسم
            """)
            rows = cur.fetchall()

            rows_data = []
            for row in rows:
                rows_data.append([
                    row["اسم_المادة"] or "",
                    row["الوحدة"] or "",
                    row["اسم_المجموعة"] or "",
                    f"{row['الكمية_المتوفرة'] or 0:.2f}",
                    str(row["آخر_تحديث"] or "")
                ])

            self.inventory_table.set_data(
                ["المادة", "الوحدة", "المجموعة", "الكمية المتوفرة", "آخر تحديث"],
                rows_data
            )
        except Exception as e:
            logger.error(str(e))
            QMessageBox.critical(self, "خطأ", f"فشل تحميل المخزون:\n{str(e)}")
        finally:
            conn.close()

    def export_inventory_csv(self):
        from PyQt5.QtWidgets import QFileDialog
        path, _ = QFileDialog.getSaveFileName(
            self, "تصدير المخزون", "inventory.csv", "CSV (*.csv)"
        )
        if not path:
            return
        conn = get_conn()
        try:
            cur = conn.cursor()
            cur.execute("""
                SELECT
                    م.الاسم AS اسم_المادة,
                    م.الوحدة,
                    ج.الاسم AS اسم_المجموعة,
                    خ.الكمية_المتوفرة,
                    خ.آخر_تحديث
                FROM المخزون خ
                JOIN المواد_الفرعية م ON خ.معرف_المادة_الفرعية = م.معرف
                JOIN المجموعات ج ON م.معرف_المجموعة = ج.معرف
                ORDER BY ج.الاسم, م.الاسم
            """)
            rows = cur.fetchall()

            import csv
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(["المادة", "الوحدة", "المجموعة", "الكمية المتوفرة", "آخر تحديث"])
                for row in rows:
                    writer.writerow([
                        row["اسم_المادة"],
                        row["الوحدة"],
                        row["اسم_المجموعة"],
                        row["الكمية_المتوفرة"],
                        row["آخر_تحديث"]
                    ])
            QMessageBox.information(self, "تم", f"تم تصدير المخزون إلى:\n{path}")
        except Exception as e:
            logger.error(str(e))
            QMessageBox.critical(self, "خطأ", f"فشل التصدير:\n{str(e)}")
        finally:
            conn.close()

    # ─────────────────────────────────────────────
    # تبويب الديون
    # ─────────────────────────────────────────────
    def build_debts_tab(self):
        layout = QVBoxLayout(self.debts_tab)
        layout.setSpacing(20)
        layout.setContentsMargins(20, 20, 20, 20)

        self.debts_summary_label = QLabel("")
        self.debts_summary_label.setStyleSheet(summary_label_style(bg="#fef9e7", border_color="#f9e79f"))
        self.debts_summary_label.setAlignment(Qt.AlignRight)
        layout.addWidget(self.debts_summary_label)

        refresh_btn = QPushButton("تحديث")
        refresh_btn.setIcon(QApplication.style().standardIcon(QStyle.SP_BrowserReload))
        refresh_btn.setCursor(Qt.PointingHandCursor)
        refresh_btn.setStyleSheet(self._btn_style("#9b59b6", "#8e44ad"))
        refresh_btn.clicked.connect(lambda: self._run_with_loading(refresh_btn, self.load_debts))
        layout.addWidget(refresh_btn)

        self.debts_table = SearchableTable(show_actions=False)
        layout.addWidget(self.debts_table)

        self.load_debts()

    def load_debts(self):
        conn = get_conn()
        try:
            cur = conn.cursor()
            cur.execute("""
                SELECT اسم_الطرف, نوع_الطرف, العملة, المبلغ_الإجمالي, المبلغ_المدفوع, الرصيد
                FROM الديون
                ORDER BY اسم_الطرف
            """)
            rows = cur.fetchall()

            rows_data = []
            total_balance_syp = 0.0
            total_balance_usd = 0.0
            total_paid_syp = 0.0
            total_paid_usd = 0.0
            for row in rows:
                balance = row["الرصيد"] or 0
                paid = row["المبلغ_المدفوع"] or 0
                rows_data.append([
                    row["اسم_الطرف"] or "",
                    row["نوع_الطرف"] or "",
                    row["العملة"] or "",
                    fmt(balance),
                    fmt(paid)
                ])
                if row["العملة"] == "دولار":
                    total_balance_usd += balance
                    total_paid_usd += paid
                else:
                    total_balance_syp += balance
                    total_paid_syp += paid

            self.debts_table.set_data(
                ["الاسم", "النوع", "العملة", "الرصيد", "المدفوع"],
                rows_data
            )

            parts = []
            if total_balance_syp > 0:
                parts.append(f"ليرة: {fmt_syp(total_balance_syp)}")
            if total_balance_usd > 0:
                parts.append(f"دولار: {fmt_usd(total_balance_usd)}")
            if self.exchange_rate and total_balance_usd > 0:
                total_eq = total_balance_syp + (total_balance_usd * self.exchange_rate)
                parts.append(f"الإجمالي بالليرة: {fmt_syp(total_eq)}")

            summary = (
                f"👥 عدد الدائنين: {len(rows)} | "
                f"💰 إجمالي الديون: {' | '.join(parts) if parts else '0'} | "
                f"💵 المدفوع ليرة: {fmt_syp(total_paid_syp)} | دولار: {fmt_usd(total_paid_usd)}"
            )
            self.debts_summary_label.setText(summary)
        except Exception as e:
            logger.error(str(e))
            QMessageBox.critical(self, "خطأ", f"فشل تحميل الديون:\n{str(e)}")
        finally:
            conn.close()

    # ─────────────────────────────────────────────
    # تبويب الديون المتأخرة
    # ─────────────────────────────────────────────
    def build_overdue_tab(self):
        layout = QVBoxLayout(self.overdue_tab)
        layout.setSpacing(20)
        layout.setContentsMargins(20, 20, 20, 20)

        header = QWidget()
        h_layout = QHBoxLayout(header)
        h_layout.setContentsMargins(0, 0, 0, 0)

        refresh_btn = QPushButton("تحديث")
        refresh_btn.setIcon(QApplication.style().standardIcon(QStyle.SP_BrowserReload))
        refresh_btn.setCursor(Qt.PointingHandCursor)
        refresh_btn.setStyleSheet(self._btn_style("#e74c3c", "#c0392b"))
        refresh_btn.clicked.connect(lambda: self._run_with_loading(refresh_btn, self.load_overdue_report))
        h_layout.addWidget(refresh_btn)

        btn_excel = QPushButton("📊 تصدير Excel")
        btn_excel.setIcon(QApplication.style().standardIcon(QStyle.SP_ArrowUp))
        btn_excel.setCursor(Qt.PointingHandCursor)
        btn_excel.setStyleSheet(self._btn_style("#e74c3c", "#c0392b"))
        btn_excel.clicked.connect(lambda: self.export_table_to_excel(
            self.overdue_table, "الديون_المتأخرة.xlsx", "الديون المتأخرة",
            ["الدائن", "النوع", "العملة", "الرصيد", "تاريخ الاستحقاق", "أيام التأخر"]
        ))
        h_layout.addWidget(btn_excel)

        btn_pdf = QPushButton("📄 تصدير PDF")
        btn_pdf.setIcon(QApplication.style().standardIcon(QStyle.SP_ArrowUp))
        btn_pdf.setCursor(Qt.PointingHandCursor)
        btn_pdf.setStyleSheet(self._btn_style("#e74c3c", "#c0392b"))
        btn_pdf.clicked.connect(lambda: self.export_table_to_pdf(
            self.overdue_table, "الديون_المتأخرة.pdf", "تقرير الديون المتأخرة",
            ["الدائن", "النوع", "العملة", "الرصيد", "تاريخ الاستحقاق", "أيام التأخر"]
        ))
        h_layout.addWidget(btn_pdf)

        h_layout.addStretch()
        layout.addWidget(header)

        self.overdue_table = SearchableTable(show_actions=False)
        layout.addWidget(self.overdue_table)

        self.overdue_summary_label = QLabel("")
        self.overdue_summary_label.setStyleSheet(summary_label_style(bg="#fef9e7", border_color="#f9e79f"))
        self.overdue_summary_label.setAlignment(Qt.AlignRight)
        layout.addWidget(self.overdue_summary_label)

        self.load_overdue_report()

    def load_overdue_report(self):
        conn = get_conn()
        try:
            rows = get_overdue_debts(conn)

            rows_data = []
            total_balance_syp = 0.0
            total_balance_usd = 0.0
            for row in rows:
                name = row["اسم_الطرف"] or ""
                ctype = row["نوع_الطرف"] or ""
                currency = row["العملة"] or ""
                balance = row["الرصيد"] or 0
                due_date = row["تاريخ_استحقاق"]
                days = row["days_overdue"] or 0

                rows_data.append([
                    name,
                    ctype,
                    currency,
                    fmt(balance),
                    due_date or "—",
                    str(days)
                ])

                if currency == "دولار":
                    total_balance_usd += balance
                else:
                    total_balance_syp += balance

            self.overdue_table.set_data(
                ["الدائن", "النوع", "العملة", "الرصيد", "تاريخ الاستحقاق", "أيام التأخر"],
                rows_data
            )

            parts = []
            if total_balance_syp > 0:
                parts.append(f"ليرة: {fmt_syp(total_balance_syp)}")
            if total_balance_usd > 0:
                parts.append(f"دولار: {fmt_usd(total_balance_usd)}")
            if self.exchange_rate and total_balance_usd > 0:
                total_eq = total_balance_syp + (total_balance_usd * self.exchange_rate)
                parts.append(f"الإجمالي بالليرة: {fmt_syp(total_eq)}")

            summary = (
                f"⏰ عدد الديون المتأخرة: {len(rows)} | "
                f"💰 إجمالي الديون المتأخرة: {' | '.join(parts) if parts else '0'}"
            )
            self.overdue_summary_label.setText(summary)
        except Exception as e:
            logger.error(str(e))
            QMessageBox.critical(self, "خطأ", f"فشل تحميل الديون المتأخرة:\n{str(e)}")
        finally:
            conn.close()

    def load_exchange_rate(self):
        conn = get_conn()
        try:
            cur = conn.cursor()
            cur.execute("SELECT القيمة FROM الإعدادات WHERE المفتاح = 'سعر_صرف_الدولار'")
            result = cur.fetchone()
            if result:
                try:
                    self.exchange_rate = float(result["القيمة"])
                except (ValueError, TypeError):
                    self.exchange_rate = None
        except Exception:
            self.exchange_rate = None
        finally:
            conn.close()

    def _on_app_data_changed(self, entity_name):
        if entity_name not in {"sales", "purchases", "materials", "creditors", "cash", "expenses", "withdrawals"}:
            return
        if not self.isVisible():
            return
        self.loading_overlay.start()
        QApplication.processEvents()
        try:
            current_widget = self.tabs.currentWidget()
            if current_widget is self.sales_tab:
                self.load_sales_report()
            elif current_widget is self.profit_tab:
                self.load_profit_report()
            elif current_widget is self.inventory_tab:
                self.load_inventory()
            elif current_widget is self.buy_list_tab:
                self.load_buy_list()
            elif current_widget is self.debts_tab:
                self.load_debts()
            elif current_widget is self.cash_tab:
                self.load_cash_movements()
            elif current_widget is self.suppliers_tab:
                self.load_suppliers_report()
            elif current_widget is self.overdue_tab:
                self.load_overdue_report()
            elif current_widget is self.best_suppliers_tab:
                self.load_best_suppliers_report()
            elif current_widget is self.tax_report_tab:
                self.load_tax_report()
        finally:
            self.loading_overlay.stop()

    # ─────────────────────────────────────────────
    # تبويب الموردون
    # ─────────────────────────────────────────────
    def _load_suppliers_combo(self):
        conn = get_conn()
        try:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT معرف, اسم_الطرف FROM الديون
                WHERE نوع_الطرف = 'مورد' ORDER BY اسم_الطرف
            """)
            self.supplier_filter_combo.clear()
            self.supplier_filter_combo.addItem("جميع الموردين", None)
            for row in cursor.fetchall():
                self.supplier_filter_combo.addItem(row[1], row[0])
        finally:
            conn.close()

    def build_suppliers_tab(self):
        layout = QVBoxLayout(self.suppliers_tab)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("المورد:"))
        self.supplier_filter_combo = QComboBox()
        self.supplier_filter_combo.addItem("جميع الموردين", None)
        self.supplier_filter_combo.setMinimumWidth(200)
        filter_layout.addWidget(self.supplier_filter_combo)

        filter_layout.addWidget(QLabel("من:"))
        self.sup_date_from = QDateEdit()
        self.sup_date_from.setDate(QDate.currentDate().addMonths(-3))
        self.sup_date_from.setCalendarPopup(True)
        filter_layout.addWidget(self.sup_date_from)

        filter_layout.addWidget(QLabel("إلى:"))
        self.sup_date_to = QDateEdit()
        self.sup_date_to.setDate(QDate.currentDate())
        self.sup_date_to.setCalendarPopup(True)
        filter_layout.addWidget(self.sup_date_to)

        view_btn = QPushButton("عرض")
        view_btn.clicked.connect(self.load_suppliers_report)
        filter_layout.addWidget(view_btn)

        export_excel_btn = QPushButton("📊 تصدير Excel")
        export_excel_btn.clicked.connect(lambda: self.export_suppliers_excel())
        filter_layout.addWidget(export_excel_btn)

        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        self.sup_summary_widget = QWidget()
        sup_summary_layout = QHBoxLayout(self.sup_summary_widget)

        self.sup_lbl_total_invoices = QLabel("عدد الفواتير: -")
        self.sup_lbl_total_amount = QLabel("إجمالي المشتريات: -")
        self.sup_lbl_total_debt = QLabel("الدين المتبقي: -")
        self.sup_lbl_total_paid = QLabel("إجمالي المدفوع: -")

        for lbl in [self.sup_lbl_total_invoices, self.sup_lbl_total_amount,
                    self.sup_lbl_total_debt, self.sup_lbl_total_paid]:
            lbl.setStyleSheet("""
                font-size: 14px; font-weight: bold; color: white;
                background-color: #2980b9; border-radius: 8px; padding: 10px 15px;
            """)
            sup_summary_layout.addWidget(lbl)

        layout.addWidget(self.sup_summary_widget)

        layout.addWidget(QLabel("📋 فواتير الشراء:"))
        self.suppliers_invoices_table = SearchableTable(show_actions=False)
        layout.addWidget(self.suppliers_invoices_table)

        layout.addWidget(QLabel("💳 حركات الديون:"))
        self.suppliers_debts_table = QTableWidget()
        self.suppliers_debts_table.setColumnCount(4)
        self.suppliers_debts_table.setHorizontalHeaderLabels([
            "التاريخ", "نوع الحركة", "المبلغ", "ملاحظات"
        ])
        self.suppliers_debts_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.suppliers_debts_table.setAlternatingRowColors(True)
        layout.addWidget(self.suppliers_debts_table)

    def build_comparison_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        filter_group = QGroupBox("📅 اختر الفترتين للمقارنة")
        fg_layout = QGridLayout(filter_group)

        fg_layout.addWidget(QLabel("الفترة الأولى - من:"), 0, 0)
        self.cmp_p1_from = QDateEdit()
        self.cmp_p1_from.setDate(QDate.currentDate().addMonths(-2))
        self.cmp_p1_from.setCalendarPopup(True)
        fg_layout.addWidget(self.cmp_p1_from, 0, 1)

        fg_layout.addWidget(QLabel("إلى:"), 0, 2)
        self.cmp_p1_to = QDateEdit()
        self.cmp_p1_to.setDate(QDate.currentDate().addMonths(-1))
        self.cmp_p1_to.setCalendarPopup(True)
        fg_layout.addWidget(self.cmp_p1_to, 0, 3)

        fg_layout.addWidget(QLabel("الفترة الثانية - من:"), 1, 0)
        self.cmp_p2_from = QDateEdit()
        self.cmp_p2_from.setDate(QDate.currentDate().addMonths(-1))
        self.cmp_p2_from.setCalendarPopup(True)
        fg_layout.addWidget(self.cmp_p2_from, 1, 1)

        fg_layout.addWidget(QLabel("إلى:"), 1, 2)
        self.cmp_p2_to = QDateEdit()
        self.cmp_p2_to.setDate(QDate.currentDate())
        self.cmp_p2_to.setCalendarPopup(True)
        fg_layout.addWidget(self.cmp_p2_to, 1, 3)

        compare_btn = QPushButton("🔍 مقارنة")
        compare_btn.setStyleSheet(self._btn_style("#2980b9", "#2471a3"))
        compare_btn.clicked.connect(self.load_comparison_report)
        fg_layout.addWidget(compare_btn, 0, 4, 2, 1)

        export_btn = QPushButton("📊 Excel")
        export_btn.setStyleSheet(self._btn_style("#27ae60", "#229954"))
        export_btn.clicked.connect(lambda: self.export_table_to_excel(
            self.comparison_table, "مقارنة_الفترات.xlsx", "المقارنة",
            ["البيان", "الفترة الأولى", "الفترة الثانية", "الفرق", "نسبة التغيير"]
        ))
        fg_layout.addWidget(export_btn, 0, 5, 2, 1)

        layout.addWidget(filter_group)

        cards_widget = QWidget()
        cards_layout = QHBoxLayout(cards_widget)

        self.cmp_cards = {}
        card_data = [
            ("المبيعات", "#27ae60"),
            ("المصروفات", "#e74c3c"),
            ("المشتريات", "#e67e22"),
            ("صافي النقدية", "#2980b9"),
        ]
        for title, color in card_data:
            card = QLabel(f"{title}\nالفترة 1: -\nالفترة 2: -\nالتغيير: -")
            card.setStyleSheet(f"""
                font-size: 13px; font-weight: bold; color: white;
                background-color: {color}; border-radius: 10px; padding: 12px;
            """)
            card.setAlignment(Qt.AlignCenter)
            cards_layout.addWidget(card)
            self.cmp_cards[title] = card

        layout.addWidget(cards_widget)

        self.comparison_table = QTableWidget()
        self.comparison_table.setColumnCount(5)
        self.comparison_table.setHorizontalHeaderLabels([
            "البيان", "الفترة الأولى", "الفترة الثانية", "الفرق", "نسبة التغيير %"
        ])
        self.comparison_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.comparison_table.horizontalHeader().setLayoutDirection(Qt.RightToLeft)
        self.comparison_table.setStyleSheet(self._table_style("#2980b9"))
        self.comparison_table.setAlternatingRowColors(True)
        layout.addWidget(self.comparison_table)

        return widget

    def load_comparison_report(self):
        p1_from = self.cmp_p1_from.date().toString("yyyy-MM-dd")
        p1_to = self.cmp_p1_to.date().toString("yyyy-MM-dd")
        p2_from = self.cmp_p2_from.date().toString("yyyy-MM-dd")
        p2_to = self.cmp_p2_to.date().toString("yyyy-MM-dd")

        conn = get_conn()
        try:
            cur = conn.cursor()

            def get_period_data(date_from, date_to):
                cur.execute("""
                    SELECT COALESCE(SUM(المبلغ_الإجمالي), 0)
                    FROM المبيعات_اليومية
                    WHERE date(normalize_date(التاريخ)) BETWEEN ? AND ?
                    AND معرف_المجموعة NOT IN (
                        SELECT معرف FROM المجموعات WHERE الاسم = 'مبيعات غير مسجلة'
                    )
                """, (date_from, date_to))
                sales = cur.fetchone()[0]

                cur.execute("""
                    SELECT COALESCE(SUM(المبلغ), 0) FROM المصروفات
                    WHERE date(normalize_date(التاريخ)) BETWEEN ? AND ?
                """, (date_from, date_to))
                expenses = cur.fetchone()[0]

                cur.execute("""
                    SELECT COALESCE(SUM(المبلغ_الإجمالي), 0)
                    FROM فواتير_الشراء
                    WHERE date(normalize_date(التاريخ)) BETWEEN ? AND ?
                """, (date_from, date_to))
                purchases = cur.fetchone()[0]

                cur.execute("""
                    SELECT COALESCE(SUM(المبلغ), 0) FROM السحوبات
                    WHERE date(normalize_date(التاريخ)) BETWEEN ? AND ?
                """, (date_from, date_to))
                withdrawals = cur.fetchone()[0]

                cur.execute("""
                    SELECT ج.الاسم, COALESCE(SUM(م.المبلغ_الإجمالي), 0)
                    FROM المبيعات_اليومية م
                    JOIN المجموعات ج ON م.معرف_المجموعة = ج.معرف
                    WHERE date(normalize_date(م.التاريخ)) BETWEEN ? AND ?
                    AND ج.الاسم != 'مبيعات غير مسجلة'
                    GROUP BY ج.معرف, ج.الاسم
                    ORDER BY ج.الاسم
                """, (date_from, date_to))
                sales_by_group = cur.fetchall()

                return {
                    "sales": sales,
                    "expenses": expenses,
                    "purchases": purchases,
                    "withdrawals": withdrawals,
                    "net": sales - expenses - purchases - withdrawals,
                    "sales_by_group": {r[0]: r[1] for r in sales_by_group}
                }

            p1 = get_period_data(p1_from, p1_to)
            p2 = get_period_data(p2_from, p2_to)

            def diff_pct(v1, v2):
                if v1 == 0:
                    return "—" if v2 == 0 else "+100%"
                pct = ((v2 - v1) / abs(v1)) * 100
                sign = "+" if pct >= 0 else ""
                return f"{sign}{pct:.1f}%"

            def diff_val(v1, v2):
                d = v2 - v1
                sign = "+" if d >= 0 else ""
                return f"{sign}{fmt(d)}"

            rows = [
                ("💰 إجمالي المبيعات", p1["sales"], p2["sales"]),
                ("💸 إجمالي المصروفات", p1["expenses"], p2["expenses"]),
                ("📦 إجمالي المشتريات", p1["purchases"], p2["purchases"]),
                ("🏦 السحوبات", p1["withdrawals"], p2["withdrawals"]),
                ("📊 صافي النقدية", p1["net"], p2["net"]),
            ]

            all_groups = set(p1["sales_by_group"]) | set(p2["sales_by_group"])
            for g in sorted(all_groups):
                v1 = p1["sales_by_group"].get(g, 0)
                v2 = p2["sales_by_group"].get(g, 0)
                rows.append((f"  ↳ {g}", v1, v2))

            self.comparison_table.setRowCount(len(rows))
            for r, (label, v1, v2) in enumerate(rows):
                self.comparison_table.setItem(r, 0, QTableWidgetItem(label))
                self.comparison_table.setItem(r, 1, QTableWidgetItem(fmt(v1)))
                self.comparison_table.setItem(r, 2, QTableWidgetItem(fmt(v2)))

                diff_item = QTableWidgetItem(diff_val(v1, v2))
                diff_item.setForeground(
                    QColor("#27ae60") if v2 >= v1 else QColor("#e74c3c")
                )
                self.comparison_table.setItem(r, 3, diff_item)

                pct_item = QTableWidgetItem(diff_pct(v1, v2))
                pct_item.setForeground(
                    QColor("#27ae60") if v2 >= v1 else QColor("#e74c3c")
                )
                self.comparison_table.setItem(r, 4, pct_item)

            for title, key in [("المبيعات", "sales"), ("المصروفات", "expenses"),
                                ("المشتريات", "purchases"), ("صافي النقدية", "net")]:
                v1, v2 = p1[key], p2[key]
                self.cmp_cards[title].setText(
                    f"{title}\n"
                    f"الفترة 1: {fmt(v1)}\n"
                    f"الفترة 2: {fmt(v2)}\n"
                    f"التغيير: {diff_val(v1, v2)} ({diff_pct(v1, v2)})"
                )

        except Exception as e:
            logger.error(str(e))
            QMessageBox.critical(self, "خطأ", str(e))
        finally:
            conn.close()

    def load_suppliers_report(self):
        selected_supplier_id = self.supplier_filter_combo.currentData()
        date_from = self.sup_date_from.date().toString("yyyy-MM-dd")
        date_to = self.sup_date_to.date().toString("yyyy-MM-dd")

        conn = get_conn()
        try:
            cursor = conn.cursor()

            if selected_supplier_id:
                cursor.execute("""
                    SELECT ف.معرف, ف.التاريخ, د.اسم_الطرف, ف.المبلغ_الإجمالي, ف.العملة
                    FROM فواتير_الشراء ف
                    JOIN الديون د ON ف.معرف_المورد = د.معرف
                    WHERE ف.معرف_المورد = ?
                    AND date(ف.التاريخ) BETWEEN ? AND ?
                    ORDER BY ف.التاريخ DESC
                """, (selected_supplier_id, date_from, date_to))
            else:
                cursor.execute("""
                    SELECT ف.معرف, ف.التاريخ, د.اسم_الطرف, ف.المبلغ_الإجمالي, ف.العملة
                    FROM فواتير_الشراء ف
                    JOIN الديون د ON ف.معرف_المورد = د.معرف
                    WHERE date(ف.التاريخ) BETWEEN ? AND ?
                    ORDER BY ف.التاريخ DESC
                """, (date_from, date_to))

            invoices = cursor.fetchall()
            invoices_data = []
            total_amount = 0.0
            for row in invoices:
                invoices_data.append([str(val or "") for val in row])
                total_amount += float(row[3] or 0)

            self.suppliers_invoices_table.set_data(
                ["رقم الفاتورة", "التاريخ", "المورد", "المبلغ الإجمالي", "العملة"],
                invoices_data
            )

            if selected_supplier_id:
                cursor.execute("""
                    SELECT ح.التاريخ, ح.نوع_الحركة, ح.المبلغ, ح.ملاحظات
                    FROM تحركات_الديون ح
                    WHERE ح.معرف_الدين = ?
                    AND date(ح.التاريخ) BETWEEN ? AND ?
                    ORDER BY ح.التاريخ DESC
                """, (selected_supplier_id, date_from, date_to))
            else:
                cursor.execute("""
                    SELECT ح.التاريخ, ح.نوع_الحركة, ح.المبلغ, ح.ملاحظات
                    FROM تحركات_الديون ح
                    JOIN الديون د ON ح.معرف_الدين = د.معرف
                    WHERE د.نوع_الطرف = 'مورد'
                    AND date(ح.التاريخ) BETWEEN ? AND ?
                    ORDER BY ح.التاريخ DESC
                """, (date_from, date_to))

            movements = cursor.fetchall()
            self.suppliers_debts_table.setRowCount(len(movements))
            for r, row in enumerate(movements):
                for c, val in enumerate(row):
                    item = QTableWidgetItem(str(val or ""))
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    if c == 1:
                        item.setForeground(
                            QColor("#e74c3c") if val == "إضافة" else QColor("#27ae60")
                        )
                    self.suppliers_debts_table.setItem(r, c, item)

            if selected_supplier_id:
                cursor.execute("""
                    SELECT المبلغ_الإجمالي, المبلغ_المدفوع, الرصيد
                    FROM الديون WHERE معرف = ?
                """, (selected_supplier_id,))
                debt_row = cursor.fetchone()
                if debt_row:
                    self.sup_lbl_total_debt.setText(f"الدين المتبقي: {debt_row[2]:,.0f}")
                    self.sup_lbl_total_paid.setText(f"إجمالي المدفوع: {debt_row[1]:,.0f}")
            else:
                cursor.execute("""
                    SELECT SUM(الرصيد), SUM(المبلغ_المدفوع)
                    FROM الديون WHERE نوع_الطرف = 'مورد'
                """)
                totals = cursor.fetchone()
                self.sup_lbl_total_debt.setText(f"الدين المتبقي: {totals[0] or 0:,.0f}")
                self.sup_lbl_total_paid.setText(f"إجمالي المدفوع: {totals[1] or 0:,.0f}")

            self.sup_lbl_total_invoices.setText(f"عدد الفواتير: {len(invoices)}")
            self.sup_lbl_total_amount.setText(f"إجمالي المشتريات: {total_amount:,.0f}")

        except Exception as e:
            logger.error(str(e))
            QMessageBox.critical(self, "خطأ", str(e))
        finally:
            conn.close()

    def export_suppliers_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from PyQt5.QtWidgets import QFileDialog
        except ImportError:
            QMessageBox.critical(self, "خطأ", "مكتبة openpyxl غير مثبتة. شغّل: pip install openpyxl")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "حفظ ملف Excel", "تقرير_الموردين.xlsx", "Excel Files (*.xlsx)"
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()

            ws1 = wb.active
            ws1.title = "فواتير الشراء"
            headers = ["رقم الفاتورة", "التاريخ", "المورد", "المبلغ الإجمالي", "العملة"]
            ws1.append(headers)
            for cell in ws1[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2980B9")
            invoices_table = self.suppliers_invoices_table.table if hasattr(self.suppliers_invoices_table, 'table') else self.suppliers_invoices_table
            for r in range(invoices_table.rowCount()):
                row_data = []
                for c in range(len(headers)):
                    item = invoices_table.item(r, c)
                    row_data.append(item.text() if item else "")
                ws1.append(row_data)

            ws2 = wb.create_sheet("حركات الديون")
            headers2 = ["التاريخ", "نوع الحركة", "المبلغ", "ملاحظات"]
            ws2.append(headers2)
            for cell in ws2[1]:
                cell.fill = PatternFill("solid", fgColor="27AE60")
                cell.font = Font(bold=True, color="FFFFFF")
            debts_table = self.suppliers_debts_table.table if hasattr(self.suppliers_debts_table, 'table') else self.suppliers_debts_table
            for r in range(debts_table.rowCount()):
                row_data = []
                for c in range(len(headers2)):
                    item = debts_table.item(r, c)
                    row_data.append(item.text() if item else "")
                ws2.append(row_data)

            ws3 = wb.create_sheet("الملخص")
            ws3.append(["البيان", "القيمة"])
            ws3.append(["عدد الفواتير", self.sup_lbl_total_invoices.text()])
            ws3.append(["إجمالي المشتريات", self.sup_lbl_total_amount.text()])
            ws3.append(["الدين المتبقي", self.sup_lbl_total_debt.text()])
            ws3.append(["إجمالي المدفوع", self.sup_lbl_total_paid.text()])

            wb.save(path)
            QMessageBox.information(self, "تم", f"تم تصدير الملف بنجاح:\n{path}")

        except Exception as e:
            logger.error(str(e))
            QMessageBox.critical(self, "خطأ", f"فشل التصدير:\n{str(e)}")

    def export_table_to_excel(self, table_widget, filename, sheet_name, headers):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from PyQt5.QtWidgets import QFileDialog
        except ImportError:
            QMessageBox.critical(self, "خطأ", "pip install openpyxl")
            return

        actual_table = table_widget.table if hasattr(table_widget, 'table') else table_widget

        path, _ = QFileDialog.getSaveFileName(
            self, "حفظ Excel", filename, "Excel Files (*.xlsx)"
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor="2C3E50")
                cell.font = Font(bold=True, color="FFFFFF")
            for r in range(actual_table.rowCount()):
                row_data = []
                for c in range(len(headers)):
                    item = actual_table.item(r, c)
                    row_data.append(item.text() if item else "")
                ws.append(row_data)
            wb.save(path)
            QMessageBox.information(self, "تم", f"تم الحفظ:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "خطأ", str(e))

    def export_table_to_pdf(self, table_widget, filename, title, headers):
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from PyQt5.QtWidgets import QFileDialog
        except ImportError:
            QMessageBox.critical(self, "خطأ", "pip install reportlab")
            return

        actual_table = table_widget.table if hasattr(table_widget, 'table') else table_widget

        path, _ = QFileDialog.getSaveFileName(
            self, "حفظ PDF", filename, "PDF Files (*.pdf)"
        )
        if not path:
            return

        try:
            doc = SimpleDocTemplate(path, pagesize=landscape(A4))
            data = [headers]
            for r in range(actual_table.rowCount()):
                row_data = []
                for c in range(len(headers)):
                    item = actual_table.item(r, c)
                    row_data.append(item.text() if item else "")
                data.append(row_data)

            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1),
                 [colors.HexColor('#FFFFFF'), colors.HexColor('#F8F9FA')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ]))

            doc.build([table])
            QMessageBox.information(self, "تم", f"تم الحفظ:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "خطأ", str(e))

    # ─────────────────────────────────────────────
    # تبويب التقرير الضريبي
    # ─────────────────────────────────────────────
    def build_tax_report_tab(self):
        layout = QVBoxLayout(self.tax_report_tab)
        layout.setSpacing(20)
        layout.setContentsMargins(20, 20, 20, 20)

        filters = QWidget()
        f_layout = QHBoxLayout(filters)
        f_layout.setContentsMargins(0, 0, 0, 0)

        f_layout.addWidget(QLabel("📅 من:"))
        self.tax_from = QDateEdit()
        self.tax_from.setDate(QDate.currentDate().addMonths(-1))
        self.tax_from.setCalendarPopup(True)
        self.tax_from.setStyleSheet(self._date_style())
        f_layout.addWidget(self.tax_from)

        f_layout.addWidget(QLabel("📅 إلى:"))
        self.tax_to = QDateEdit()
        self.tax_to.setDate(QDate.currentDate())
        self.tax_to.setCalendarPopup(True)
        self.tax_to.setStyleSheet(self._date_style())
        f_layout.addWidget(self.tax_to)

        show_btn = QPushButton("عرض")
        show_btn.setIcon(QApplication.style().standardIcon(QStyle.SP_FileDialogDetailedView))
        show_btn.setCursor(Qt.PointingHandCursor)
        show_btn.setStyleSheet(self._btn_style("#8e44ad", "#7f4aa0"))
        show_btn.clicked.connect(lambda: self._run_with_loading(show_btn, self.load_tax_report))
        f_layout.addWidget(show_btn)

        btn_excel = QPushButton("📊 تصدير Excel")
        btn_excel.setIcon(QApplication.style().standardIcon(QStyle.SP_ArrowUp))
        btn_excel.setCursor(Qt.PointingHandCursor)
        btn_excel.setStyleSheet(self._btn_style("#8e44ad", "#7f4aa0"))
        btn_excel.clicked.connect(lambda: self.export_table_to_excel(
            self.tax_table, "التقرير_الضريبي.xlsx", "التقرير الضريبي",
            ["البيان", "ليرة سورية", "دولار"]
        ))
        f_layout.addWidget(btn_excel)

        btn_pdf = QPushButton("📄 تصدير PDF")
        btn_pdf.setIcon(QApplication.style().standardIcon(QStyle.SP_ArrowUp))
        btn_pdf.setCursor(Qt.PointingHandCursor)
        btn_pdf.setStyleSheet(self._btn_style("#8e44ad", "#7f4aa0"))
        btn_pdf.clicked.connect(lambda: self.export_table_to_pdf(
            self.tax_table, "التقرير_الضريبي.pdf", "التقرير الضريبي",
            ["البيان", "ليرة سورية", "دولار"]
        ))
        f_layout.addWidget(btn_pdf)

        f_layout.addStretch()
        layout.addWidget(filters)

        self.tax_table = QTableWidget()
        self.tax_table.setColumnCount(3)
        self.tax_table.setHorizontalHeaderLabels([
            "البيان", "ليرة سورية", "دولار"
        ])
        self.tax_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        for col in range(1, 3):
            self.tax_table.horizontalHeader().setSectionResizeMode(col, QHeaderView.ResizeToContents)
        self.tax_table.horizontalHeader().setLayoutDirection(Qt.RightToLeft)
        self.tax_table.setStyleSheet(self._table_style("#8e44ad"))
        self.tax_table.setAlternatingRowColors(True)
        layout.addWidget(self.tax_table)

        self.tax_summary_label = QLabel("")
        self.tax_summary_label.setStyleSheet(summary_label_style(bg="#fef9e7", border_color="#f9e79f"))
        self.tax_summary_label.setAlignment(Qt.AlignRight)
        self.tax_summary_label.setWordWrap(True)
        layout.addWidget(self.tax_summary_label)

    def load_tax_report(self):
        from_date = self.tax_from.date().toString("yyyy-MM-dd")
        to_date = self.tax_to.date().toString("yyyy-MM-dd")
        from_date = from_date.translate(str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789"))
        to_date = to_date.translate(str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789"))
        conn = get_conn()
        try:
            cur = conn.cursor()

            cur.execute("""
                SELECT العملة, COALESCE(SUM(المبلغ_الإجمالي), 0) AS الإجمالي
                FROM المبيعات_اليومية
                WHERE date(normalize_date(التاريخ)) >= ? AND date(normalize_date(التاريخ)) <= ?
                AND معرف_المجموعة NOT IN (
                    SELECT معرف FROM المجموعات WHERE الاسم = 'مبيعات غير مسجلة'
                )
                GROUP BY العملة
            """, (from_date, to_date))
            sales_rows = cur.fetchall()

            cur.execute("""
                SELECT العملة, COALESCE(SUM(المبلغ_الإجمالي), 0) AS الإجمالي
                FROM فواتير_الشراء
                WHERE date(normalize_date(التاريخ)) >= ? AND date(normalize_date(التاريخ)) <= ?
                GROUP BY العملة
            """, (from_date, to_date))
            purchase_rows = cur.fetchall()

            sales_syp = sum(r["الإجمالي"] for r in sales_rows if r["العملة"] == "ليرة_سورية")
            sales_usd = sum(r["الإجمالي"] for r in sales_rows if r["العملة"] == "دولار")
            purchases_syp = sum(r["الإجمالي"] for r in purchase_rows if r["العملة"] == "ليرة_سورية")
            purchases_usd = sum(r["الإجمالي"] for r in purchase_rows if r["العملة"] == "دولار")

            diff_syp = sales_syp - purchases_syp
            diff_usd = sales_usd - purchases_usd

            rows_data = [
                ["إجمالي المبيعات", fmt(sales_syp), fmt(sales_usd)],
                ["إجمالي المشتريات", fmt(purchases_syp), fmt(purchases_usd)],
                ["الفرق (مبيعات − مشتريات)", fmt(diff_syp), fmt(diff_usd)],
            ]

            self.tax_table.setRowCount(len(rows_data))
            for r, row in enumerate(rows_data):
                for c, val in enumerate(row):
                    item = QTableWidgetItem(str(val))
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    self.tax_table.setItem(r, c, item)

            parts = []
            if sales_syp or purchases_syp:
                parts.append(f"ليرة: مبيعات {fmt_syp(sales_syp)} | مشتريات {fmt_syp(purchases_syp)}")
            if sales_usd or purchases_usd:
                parts.append(f"دولار: مبيعات {fmt_usd(sales_usd)} | مشتريات {fmt_usd(purchases_usd)}")

            summary = " | ".join(parts) if parts else "ليرة: 0 | دولار: 0"

            if self.exchange_rate and (sales_usd or purchases_usd):
                converted_diff = diff_syp + (diff_usd * self.exchange_rate)
                summary += f" | التحويل الاختياري إلى ليرة: {fmt_syp(converted_diff)}"
                summary += "\nⓘ يستخدم سعر الصرف الحالي، وليس الأسعار التاريخية لكل معاملة."

            warning = (
                "تنبيه مهم: هذا الرقم ليس ربحاً صافياً حقيقياً ولا يخصم "
                "المصروفات أو الرواتب، ولا يعكس تكلفة البضاعة المباعة الفعلية "
                "التي تتطلب جرداً دورياً كما هو موثق في README.md قسم 7.5."
            )
            summary += f"\n⚠️ {warning}"

            self.tax_summary_label.setText(summary)
        except Exception as e:
            logger.error(str(e))
            QMessageBox.critical(self, "خطأ", f"فشل تحميل التقرير الضريبي:\n{str(e)}")
        finally:
            conn.close()

    # ─────────────────────────────────────────────
    # تبويب أفضل الموردين
    # ─────────────────────────────────────────────
    def build_best_suppliers_tab(self):
        layout = QVBoxLayout(self.best_suppliers_tab)
        layout.setSpacing(20)
        layout.setContentsMargins(20, 20, 20, 20)

        header = QWidget()
        h_layout = QHBoxLayout(header)
        h_layout.setContentsMargins(0, 0, 0, 0)

        subtitle = QLabel("📌 مقارنة موردين مختلفين لنفس المادة (لا علاقة بمقارنة أسعار مورد واحد عبر الزمن)")
        subtitle.setStyleSheet(f"font-size: {FontSizes.MD}; color: {Colors.GRAY}; font-style: italic;")
        h_layout.addWidget(subtitle)

        refresh_btn = QPushButton("تحديث")
        refresh_btn.setIcon(QApplication.style().standardIcon(QStyle.SP_BrowserReload))
        refresh_btn.setCursor(Qt.PointingHandCursor)
        refresh_btn.setStyleSheet(self._btn_style("#27ae60", "#229954"))
        refresh_btn.clicked.connect(lambda: self._run_with_loading(refresh_btn, self.load_best_suppliers_report))
        h_layout.addWidget(refresh_btn)

        btn_excel = QPushButton("📊 تصدير Excel")
        btn_excel.setIcon(QApplication.style().standardIcon(QStyle.SP_ArrowUp))
        btn_excel.setCursor(Qt.PointingHandCursor)
        btn_excel.setStyleSheet(self._btn_style("#27ae60", "#229954"))
        btn_excel.clicked.connect(lambda: self.export_table_to_excel(
            self.best_suppliers_table, "أفضل_الموردين.xlsx", "أفضل الموردين",
            ["المادة", "المورد", "آخر سعر"]
        ))
        h_layout.addWidget(btn_excel)

        btn_pdf = QPushButton("📄 تصدير PDF")
        btn_pdf.setIcon(QApplication.style().standardIcon(QStyle.SP_ArrowUp))
        btn_pdf.setCursor(Qt.PointingHandCursor)
        btn_pdf.setStyleSheet(self._btn_style("#e74c3c", "#c0392b"))
        btn_pdf.clicked.connect(lambda: self.export_table_to_pdf(
            self.best_suppliers_table, "أفضل_الموردين.pdf", "تقرير أفضل الموردين",
            ["المادة", "المورد", "آخر سعر"]
        ))
        h_layout.addWidget(btn_pdf)

        h_layout.addStretch()
        layout.addWidget(header)

        self.best_suppliers_table = SearchableTable(show_actions=False)
        layout.addWidget(self.best_suppliers_table)

        self.best_suppliers_info = QLabel("")
        self.best_suppliers_info.setStyleSheet(summary_label_style())
        self.best_suppliers_info.setAlignment(Qt.AlignRight)
        layout.addWidget(self.best_suppliers_info)

    def load_best_suppliers_report(self):
        conn = get_conn()
        try:
            cur = conn.cursor()
            cur.execute("""
                WITH latest_invoice_dates AS (
                    SELECT
                        d2.معرف_المادة_الفرعية,
                        f2.معرف_المورد,
                        MAX(f2.التاريخ) AS احدث_تاريخ
                    FROM فواتير_الشراء f2
                    JOIN تفاصيل_الشراء d2 ON f2.معرف = d2.معرف_الفاتورة
                    JOIN الديون d3 ON f2.معرف_المورد = d3.معرف
                    WHERE d3.نوع_الطرف = 'مورد'
                    GROUP BY d2.معرف_المادة_الفرعية, f2.معرف_المورد
                ),
                latest_invoices AS (
                    SELECT
                        d2.معرف_المادة_الفرعية,
                        f2.معرف_المورد,
                        MAX(f2.معرف) AS معرف_الفاتورة_الأخيرة
                    FROM فواتير_الشراء f2
                    JOIN تفاصيل_الشراء d2 ON f2.معرف = d2.معرف_الفاتورة
                    JOIN latest_invoice_dates lid ON
                        d2.معرف_المادة_الفرعية = lid.معرف_المادة_الفرعية
                        AND f2.معرف_المورد = lid.معرف_المورد
                        AND f2.التاريخ = lid.احدث_تاريخ
                    GROUP BY d2.معرف_المادة_الفرعية, f2.معرف_المورد
                )
                SELECT
                    م.الاسم AS اسم_المادة,
                    د.اسم_الطرف AS اسم_المورد,
                    تف.سعر_الوحدة AS سعر_الوحدة
                FROM latest_invoices li
                JOIN تفاصيل_الشراء تف ON li.معرف_الفاتورة_الأخيرة = تف.معرف_الفاتورة
                    AND li.معرف_المادة_الفرعية = تف.معرف_المادة_الفرعية
                JOIN المواد_الفرعية م ON تف.معرف_المادة_الفرعية = م.معرف
                JOIN الديون د ON li.معرف_المورد = د.معرف
                ORDER BY م.الاسم, تف.سعر_الوحدة ASC
            """)
            rows = cur.fetchall()

            from collections import defaultdict
            material_suppliers = defaultdict(list)
            for row in rows:
                material_suppliers[row["اسم_المادة"]].append(row)

            filtered_rows = []
            for material_name, suppliers in material_suppliers.items():
                if len(suppliers) > 1:
                    suppliers_sorted = sorted(suppliers, key=lambda x: x["سعر_الوحدة"] or 0)
                    filtered_rows.extend(suppliers_sorted)

            rows_data = []
            for row in filtered_rows:
                rows_data.append([
                    row["اسم_المادة"] or "",
                    row["اسم_المورد"] or "",
                    fmt(row["سعر_الوحدة"] or 0)
                ])

            self.best_suppliers_table.set_data(
                ["المادة", "المورد", "آخر سعر"],
                rows_data
            )

            materials_count = len(material_suppliers)
            multi_supplier_count = sum(1 for s in material_suppliers.values() if len(s) > 1)
            single_supplier_count = materials_count - multi_supplier_count

            if not filtered_rows:
                self.best_suppliers_info.setText(
                    "📋 لا توجد مواد مُشتَّرى من أكثر من مورد واحد."
                )
            else:
                self.best_suppliers_info.setText(
                    f"📊 المواد المقارَنة: {multi_supplier_count} | "
                    f"ℹ️ المواد بمورد واحد (غير مُعرَّضة): {single_supplier_count}"
                )
        except Exception as e:
            logger.error(str(e))
            QMessageBox.critical(self, "خطأ", f"فشل تحميل تقرير أفضل الموردين:\n{str(e)}")
        finally:
            conn.close()

    # ─────────────────────────────────────────────
    # أنماط مساعدة
    # ─────────────────────────────────────────────
    # تبويب حركة النقدية
    # ─────────────────────────────────────────────
    def build_cash_tab(self):
        layout = QVBoxLayout(self.cash_tab)
        layout.setSpacing(20)
        layout.setContentsMargins(20, 20, 20, 20)

        filters = QWidget()
        f_layout = QHBoxLayout(filters)
        f_layout.setContentsMargins(0, 0, 0, 0)

        f_layout.addWidget(QLabel("📅 من:"))
        self.cash_from = QDateEdit()
        self.cash_from.setDate(QDate.currentDate().addMonths(-1))
        self.cash_from.setCalendarPopup(True)
        self.cash_from.setStyleSheet(self._date_style())
        f_layout.addWidget(self.cash_from)

        f_layout.addWidget(QLabel("📅 إلى:"))
        self.cash_to = QDateEdit()
        self.cash_to.setDate(QDate.currentDate())
        self.cash_to.setCalendarPopup(True)
        self.cash_to.setStyleSheet(self._date_style())
        f_layout.addWidget(self.cash_to)

        show_btn = QPushButton("عرض الحركة")
        show_btn.setIcon(QApplication.style().standardIcon(QStyle.SP_FileDialogDetailedView))
        show_btn.setCursor(Qt.PointingHandCursor)
        show_btn.setStyleSheet(self._btn_style("#e67e22", "#d35400"))
        show_btn.clicked.connect(lambda: self._run_with_loading(show_btn, self.load_cash_movements))
        f_layout.addWidget(show_btn)

        btn_excel = QPushButton("📊 تصدير Excel")
        btn_excel.setIcon(QApplication.style().standardIcon(QStyle.SP_ArrowUp))
        btn_excel.setCursor(Qt.PointingHandCursor)
        btn_excel.setStyleSheet(self._btn_style("#e67e22", "#d35400"))
        btn_excel.clicked.connect(lambda: self.export_table_to_excel(
            self.cash_table, "حركة_النقدية.xlsx", "النقدية",
            ["التاريخ", "بداية", "مبيعات", "مصروفات", "سحوبات", "رصيد نظري", "فعلي", "فرق", "رصيد الخزنة"]
        ))
        f_layout.addWidget(btn_excel)

        btn_pdf = QPushButton("📄 تصدير PDF")
        btn_pdf.setIcon(QApplication.style().standardIcon(QStyle.SP_ArrowUp))
        btn_pdf.setCursor(Qt.PointingHandCursor)
        btn_pdf.setStyleSheet(self._btn_style("#e67e22", "#d35400"))
        btn_pdf.clicked.connect(lambda: self.export_table_to_pdf(
            self.cash_table, "حركة_النقدية.pdf", "تقرير حركة النقدية",
            ["التاريخ", "بداية", "مبيعات", "مصروفات", "سحوبات", "رصيد نظري", "فعلي", "فرق", "رصيد الخزنة"]
        ))
        f_layout.addWidget(btn_pdf)

        f_layout.addStretch()
        layout.addWidget(filters)

        self.cash_table = QTableWidget()
        self.cash_table.setColumnCount(9)
        self.cash_table.setHorizontalHeaderLabels([
            "التاريخ", "بداية", "مبيعات", "مصروفات", "سحوبات",
            "رصيد نظري", "فعلي", "فرق", "رصيد الخزنة"
        ])
        self.cash_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        for col in range(1, 8):
            self.cash_table.horizontalHeader().setSectionResizeMode(col, QHeaderView.ResizeToContents)
        self.cash_table.horizontalHeader().setSectionResizeMode(8, QHeaderView.Stretch)
        self.cash_table.horizontalHeader().setLayoutDirection(Qt.RightToLeft)
        self.cash_table.setStyleSheet(self._table_style("#e67e22"))
        self.cash_table.setAlternatingRowColors(True)
        layout.addWidget(self.cash_table)

    def load_cash_movements(self):
        from_date = self.cash_from.date().toString("yyyy-MM-dd")
        to_date = self.cash_to.date().toString("yyyy-MM-dd")
        from_date = from_date.translate(str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789"))
        to_date = to_date.translate(str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789"))
        conn = get_conn()
        try:
            cur = conn.cursor()
            try:
                cur.execute("""
                    SELECT
                        التاريخ,
                        رصيد_بداية_اليوم,
                        مبيعات_اليوم,
                        مصروفات_اليوم,
                        سحوبات_اليوم,
                        رصيد_نهاية_نظري,
                        رصيد_نهاية_فعلي,
                        فرق_التسوية,
                        رصيد_الخزنة
                    FROM أرصدة_الصندوق
                    WHERE date(normalize_date(التاريخ)) >= ? AND date(normalize_date(التاريخ)) <= ?
                    ORDER BY التاريخ DESC
                """, (from_date, to_date))
                has_vault_col = True
            except Exception:
                cur.execute("""
                    SELECT
                        التاريخ,
                        رصيد_بداية_اليوم,
                        مبيعات_اليوم,
                        مصروفات_اليوم,
                        سحوبات_اليوم,
                        رصيد_نهاية_نظري,
                        رصيد_نهاية_فعلي,
                        فرق_التسوية
                    FROM أرصدة_الصندوق
                    WHERE date(normalize_date(التاريخ)) >= ? AND date(normalize_date(التاريخ)) <= ?
                    ORDER BY التاريخ DESC
                """, (from_date, to_date))
                has_vault_col = False
            rows = cur.fetchall()

            self.cash_table.setRowCount(len(rows))
            for r, row in enumerate(rows):
                self.cash_table.setItem(r, 0, QTableWidgetItem(str(row["التاريخ"] or "")))
                self.cash_table.setItem(r, 1, QTableWidgetItem(fmt(row['رصيد_بداية_اليوم'] or 0)))
                self.cash_table.setItem(r, 2, QTableWidgetItem(fmt(row['مبيعات_اليوم'] or 0)))
                self.cash_table.setItem(r, 3, QTableWidgetItem(fmt(row['مصروفات_اليوم'] or 0)))
                self.cash_table.setItem(r, 4, QTableWidgetItem(fmt(row['سحوبات_اليوم'] or 0)))
                self.cash_table.setItem(r, 5, QTableWidgetItem(fmt(row['رصيد_نهاية_نظري'] or 0)))
                self.cash_table.setItem(r, 6, QTableWidgetItem(fmt(row['رصيد_نهاية_فعلي'] or 0)))
                diff = row["فرق_التسوية"] or 0
                diff_item = QTableWidgetItem(fmt(diff))
                if diff > 0.01:
                    diff_item.setForeground(QColor("#27ae60"))
                elif diff < -0.01:
                    diff_item.setForeground(QColor("#e74c3c"))
                self.cash_table.setItem(r, 7, diff_item)
                vault_balance = row["رصيد_الخزنة"] if has_vault_col and "رصيد_الخزنة" in row.keys() else 0
                self.cash_table.setItem(r, 8, QTableWidgetItem(fmt(vault_balance)))
        except Exception as e:
            logger.error(str(e))
            QMessageBox.critical(self, "خطأ", f"فشل تحميل حركة النقدية:\n{str(e)}")
        finally:
            conn.close()

    # ─────────────────────────────────────────────
    # أنماط مساعدة
    # ─────────────────────────────────────────────
    def _date_style(self):
        return date_edit_style()

    def _btn_style(self, bg, hover):
        return primary_button_style(bg=bg, hover=hover, font_size=FontSizes.LG, padding="10px 20px")

    def _table_style(self, header_color):
        return table_style(header_color)
